import logging
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from copy import copy
from datetime import datetime, timedelta
import os
import time
from typing import List, Pattern, Any
import pandas as pd
from utils.utils import validate_frescures, frescure_to_date, validate_sku

logger = logging.getLogger(__name__)

# Constantes de la plantilla
ROWS_PER_PAGE = 25  # Filas que ocupa cada pagina A4
MAX_COL = 5  # Columnas A-E (5 columnas)
# Celdas donde se inyectan datos (fila relativa dentro de cada bloque)
ROW_FRESCURA = 8   # D8
ROW_SKU = 15       # D15
ROW_CADUCIDAD = 24 # D24
COL_DATA = 4       # Columna D


class Frescurer:
    def __init__(self, shelf_time_path: str, template_path: str, output_path: str, query: List[List[str]], project_root: str, frescures_pattern: Pattern[Any]):
        t0 = time.perf_counter()
        self.project_root = project_root
        self.shelf_table = self.load_data(shelf_time_path)
        self.template_path = template_path
        self.output_path = output_path
        self.frescures_pattern = frescures_pattern
        all_frescures = self.validate_query(query)
        self.attend_query(self.shelf_table, all_frescures, self.template_path)
        logger.info(f"Proceso completado en: {time.perf_counter() - t0:.6f}")

    def load_data(self, shelf_time_table: str) -> pd.DataFrame:
        try:
            DF_DIAS = pd.read_csv(shelf_time_table, encoding='utf-8') #type: ignore
            return DF_DIAS
        except FileNotFoundError as e:
            logger.error(f"Error no se encontro archivo con dias de consumo preferente: '{e}'", exc_info=True)
            return pd.DataFrame({'CODIGO': [] ,'DESCRIPCION': [], 'SHELF_LIFE': []})

    def validate_query(self, all_frescuras: List[List[str]]) -> List[List[str]]:
        complete_frescures: List[List[str]] = []
        for frescuras in all_frescuras:
            if not validate_frescures(self.frescures_pattern, frescuras[1]) or not validate_sku(frescuras[0]):
                continue
                        
            fecha = frescure_to_date(frescuras[1])
            complete_frescures.append([frescuras[0], frescuras[1], fecha])

        return complete_frescures
    
    def attend_query(self, shelf_table: pd.DataFrame, all_frescures: List[List[str]], template_path: str):
        complete_data: List[List[str]] = []
        for frescure in all_frescures:
            sku = frescure[0].strip()
            codigo_frescura = frescure[1]
            frescura_final = frescure[2]
            
            fecha_base = datetime.strptime(frescure[2], "%d/%m/%Y").date()
            
            df_match = shelf_table.loc[shelf_table['CODIGO'].astype(str) == sku, 'SHELF_LIFE']

            if df_match.empty:
                logger.warning(f"SKU {sku} no encontrado en tabla de shelf life.")
                continue

            shelf_life_days = int(df_match.iloc[0])
            fecha_consumo_preferente = fecha_base + timedelta(days=shelf_life_days)
            fecha_final_consumo = fecha_consumo_preferente.strftime("%d/%m/%Y")

            complete_data.append([frescure[0], codigo_frescura, frescura_final, fecha_final_consumo])

        logger.info(f"Final Query: {complete_data}")
        self.create_templates(complete_data, template_path)

    def create_templates(self, complete_data: List[List[str]], template_path: str):
        template = openpyxl.load_workbook(template_path)
        
        hoja = template.active
        if hoja is None:
            return
        
        # 1. Analizar plantilla original (Filas 1 a ROWS_PER_PAGE)
        
        # Guardar alturas de fila
        row_heights = {}
        for row in range(1, ROWS_PER_PAGE + 1):
            if row in hoja.row_dimensions:
                row_heights[row] = hoja.row_dimensions[row].height
        
        # Guardar celdas combinadas que estén dentro del rango de la plantilla
        merged_ranges = []
        for merged_range in hoja.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= ROWS_PER_PAGE:
                merged_ranges.append((min_col, min_row, max_col, max_row))

        # Guardar contenido y estilos celda por celda
        template_cells = {}
        for row in range(1, ROWS_PER_PAGE + 1):
            for col in range(1, MAX_COL + 1):
                cell = hoja.cell(row=row, column=col)
                template_cells[(row, col)] = {
                    'value': cell.value,
                    'font': copy(cell.font) if cell.has_style else None,
                    'border': copy(cell.border) if cell.has_style else None,
                    'fill': copy(cell.fill) if cell.has_style else None,
                    'number_format': copy(cell.number_format) if cell.has_style else None,
                    'protection': copy(cell.protection) if cell.has_style else None,
                    'alignment': copy(cell.alignment) if cell.has_style else None,
                }
        
        # 2. Generar copias
        for idx, data in enumerate(complete_data):
            sku = str(data[0])
            frescura = str(data[2])
            caducidad = str(data[3])
            
            # El primer registro (idx=0) ya usa la plantilla original en filas 1-25
            # Solo escribimos datos.
            
            row_offset = idx * ROWS_PER_PAGE
            
            # Si es una copia (idx > 0), replicamos estructura
            if idx > 0:
                # A. Copiar filas y celdas
                for row in range(1, ROWS_PER_PAGE + 1):
                    target_row = row + row_offset
                    
                    # Copiar altura
                    if row in row_heights:
                        hoja.row_dimensions[target_row].height = row_heights[row]
                    
                    # Copiar celdas
                    for col in range(1, MAX_COL + 1):
                        source_data = template_cells.get((row, col))
                        if source_data:
                            target_cell = hoja.cell(row=target_row, column=col)
                            target_cell.value = source_data['value']
                            
                            if source_data['font']: target_cell.font = copy(source_data['font'])
                            if source_data['border']: target_cell.border = copy(source_data['border'])
                            if source_data['fill']: target_cell.fill = copy(source_data['fill'])
                            if source_data['number_format']: target_cell.number_format = source_data['number_format']
                            if source_data['protection']: target_cell.protection = copy(source_data['protection'])
                            if source_data['alignment']: target_cell.alignment = copy(source_data['alignment'])
                
                # B. Replicar celdas combinadas (merged cells) con el offset
                for (min_col, min_row, max_col, max_row) in merged_ranges:
                    # Ajustar filas con el offset
                    new_min_row = min_row + row_offset
                    new_max_row = max_row + row_offset
                    
                    # Crear string del rango, ej: "A26:B28"
                    start_cell = f"{get_column_letter(min_col)}{new_min_row}"
                    end_cell = f"{get_column_letter(max_col)}{new_max_row}"
                    hoja.merge_cells(f"{start_cell}:{end_cell}")

                # C. Agregar salto de página
                hoja.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=row_offset))

            # 3. Inyectar datos
            hoja.cell(row=ROW_FRESCURA + row_offset, column=COL_DATA).value = frescura
            hoja.cell(row=ROW_SKU + row_offset, column=COL_DATA).value = sku
            hoja.cell(row=ROW_CADUCIDAD + row_offset, column=COL_DATA).value = caducidad
        
        # Guardar
        os.makedirs(self.output_path, exist_ok=True)
        excel_path = f"{self.output_path}/hojas_de_frescura.xlsx"
        template.save(excel_path)
        logger.info(f"Documento generado: {excel_path}")
